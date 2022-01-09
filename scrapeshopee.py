from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from selenium.common.exceptions import StaleElementReferenceException
from selenium.webdriver.support import expected_conditions as EC

from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.styles import colors
from openpyxl.styles import Font

from bs4 import BeautifulSoup
from time import sleep
from time import time
import pandas as pd


shopee_list = pd.read_excel('shopee_item_list.xlsx' )

chrome_options = Options()
chrome_options.add_argument('disable-notifications')
chrome_options.add_argument('--disable-infobars')
chrome_options.add_argument('start-maximized')
chrome_options.add_argument("disable-infobars")
chrome_options.add_experimental_option("prefs", {"profile.default_content_setting_values.notifications": 2})
chrome_options.add_experimental_option("excludeSwitches", ["enable-logging"])


def get_url(search_term):
    """Generate an url from the search term"""
    template = "https://www.shopee.sg/search?keyword={}"
    search_term = search_term.replace(' ', '+')
    url = template.format(search_term)
    url += '&page={}&sortBy=sales'
    return url


def main(search_term):

    driver = webdriver.Chrome(ChromeDriverManager().install() ,options=chrome_options)
    rows = []
    url = get_url(search_term)
    before = time()
    for page in range(0, 1):
        driver.get(url.format(page))
        try:
            WebDriverWait(driver, 20).until(EC.presence_of_all_elements_located((By.CLASS_NAME, "shopee-search-item-result__item")))
            sleep(5)
            html = driver.page_source
            soup = BeautifulSoup(html, "html.parser")
            for item in soup.find_all('div', {'class': 'col-xs-2-4 shopee-search-item-result__item'}):
                name = item.find('div', {'class': '_10Wbs- _2STCsK _3IqNCf'})
                if name is not None:
                    name = name.text.strip()
                else:
                    name = ''

                price = item.find('div', {'class': 'zp9xm9 kNGSLn l-u0xK'})
                if price is not None:
                    price = str(price.find('span', {'class': '_3c5u7X'}).text.strip())
                else:
                    price = ''
                print([search_term, price, name])
                rows.append([search_term, price, name])
                if len(rows) == 3:
                    break

            driver.close()
            after = time()

        except (TimeoutException, StaleElementReferenceException):
            break

    df = pd.DataFrame(rows, columns = ['Search Term', 'Top Price', 'Top Item Name'])
    df.index += 1
    path = 'shopee_item_list.xlsx'
    book = load_workbook(path)
    writer = pd.ExcelWriter(path, engine='openpyxl', options={'strings_to_formulas': False}, mode = 'w')
    writer.book = book
    fontStyle = Font(name="Calibri", size=12, color=colors.BLACK)
    if (len(df) > 0):
        writer.font = fontStyle
        df.to_excel(writer, sheet_name = "top 3 {}".format(search_term), index=True, startrow=1, startcol=0)
        print("Scraping {} complete in {}s".format(search_term, round(after-before,2)))
    elif (len(df) == 0):
        ws = book.create_sheet('top 3 {}'.format(search_term))
        ws.cell(3,2).value = 'No results for {}'.format(search_term)
        print("0 records found for {}".format(search_term))
    else:
        print("Error, unable to extract {}".format(search_term))
    writer.close()

for i in shopee_list['Search Terms']:
    main(i)
    sleep(20)